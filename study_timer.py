import random, json
from datetime import datetime
from tkinter import *
from tkinter.messagebox import askyesno
from openpyxl import load_workbook
from settings import TimeUnit

# pick a color from the color chart
with open("color_chart.json", 'r') as file:
    color_chart = json.load(file)

if random.randint(0, 1): color_hex = f"#{random.choice(color_chart['white_font'])}"
else: color_hex = f"#{random.choice(color_chart['black_font'])}"

FONT_8 = ('Arial', 8)
FONT_12 = ('Arial', 12)
FONT_16 = ('Arial', 16)
FONT_18 = ('Arial', 18, 'underline')
FONT_48 = ('Arial', 48)

VER = 'Ver 1.0.0'

# ---------------------------- UI SETUP ------------------------------- #
class StudyTimer:
    def __init__(self, time_unit: TimeUnit):
        self.time_unit = time_unit # from settings import TimeUnit
        
        self.window = Tk()
        self.window.title('Study Timer')
        self.window.config(padx=20, pady=20, width=380, height=600, bg=color_hex)

        # Label
        self.title_label = Label(text='S   T   U   D   Y      T   I   M   E   R', 
                                fg='white', font=FONT_16, bg=color_hex)
        self.title_label.grid(row=0, column=0, columnspan=3, pady=(40, 0))
        
        self.timer_label = Label(text=self.time_unit.record_formatting(self.time_unit.duration_of_the_day), 
                                fg='white', font=FONT_48, bg=color_hex)
        self.timer_label.grid(row=1, column=0, columnspan=3)
        
        self.this_month_label = Label(text='this month {0}'.format(self.time_unit.record_formatting(self.time_unit.duration_this_month)), 
                                    fg='white', font=FONT_12, bg=color_hex)
        self.this_month_label.grid(row=2, column=0, columnspan=3)
        
        self.until_now_label = Label(text='until now {0}'.format(self.time_unit.record_formatting(self.time_unit.duration_until_now)), 
                                    fg='white', font=FONT_12, bg=color_hex)
        self.until_now_label.grid(row=3, column=0, columnspan=3)

        self.version_label = Label(text=VER, fg='white', font=FONT_8, bg=color_hex)
        self.version_label.grid(row=5, column=2)
        
        # Button
        self.reset_btn = Button(text='RESET', font=FONT_18, fg='white', 
                                bg=color_hex, bd=0, activeforeground=color_hex, 
                                command=self.reset_timer, cursor='hand2')
        self.reset_btn.grid(row=4, column=0, pady=(20, 0))
        
        self.start_btn = Button(text='START', font=FONT_18, fg='white', 
                                bg=color_hex, bd=0, activeforeground=color_hex, 
                                command=self.change_start_btn, cursor='hand2')
        self.start_btn.grid(row=4, column=1, pady=(20, 0))
        
        self.save_btn = Button(text='SAVE', font=FONT_18, fg='white', 
                            bg=color_hex, bd=0, activeforeground=color_hex, 
                            command=self.save_timer, cursor='hand2')
        self.save_btn.grid(row=4, column=2, pady=(20, 0))
        
        self.change_fg('black')
        self.window.mainloop()
    
    # Change fg color of widgets (to 'black')
    def change_fg(self, color):
        if color_hex[1:] in color_chart['black_font']:
            for wdg in self.window.children:
                target_wdg = self.window.nametowidget(wdg)
                if isinstance(target_wdg, Label):
                    target_wdg.config(fg=color)
                elif isinstance(target_wdg, Button):
                    target_wdg.config(fg=color, activebackground=color)

    # Start or Stop the timer
    def change_start_btn(self):
        if self.start_btn.cget('text') == 'START':
            self.start_timer()
            
        elif self.start_btn.cget('text') == 'STOP':
            self.stop_timer()
        
    def start_timer(self):
        self.time_unit.duration_of_the_day += 1
        
        self.timer_label.config(text=self.time_unit.record_formatting(self.time_unit.duration_of_the_day))
        self.start_btn.config(text='STOP')
        self.tk_after = self.window.after(1000, self.start_timer) # recursion
        
    def stop_timer(self):
        self.window.after_cancel(self.tk_after)
        self.start_btn.config(text='START')
        
    def reset_timer(self):
        answer = askyesno(title='Reset',
                        message='Are you sure that you want to reset the timer?')
        if answer:
            try: self.stop_timer()
            except: pass
            
            self.time_unit.duration_of_the_day = 0
            self.timer_label.config(text=self.time_unit.record_formatting(self.time_unit.duration_of_the_day))

    def save_timer(self):
        try: self.stop_timer()
        except: pass
        
        answer = askyesno(title='Save',
                        message='Are you sure that you want to save te time?')
        
        if answer:        
            TODAY = datetime.now().strftime("%A").upper()[:3]
            THIS_MONTH = datetime.now().strftime("%B").upper()[0:3]
            THIS_YEAR = datetime.now().strftime("%Y")
                        
            wb = load_workbook('성공2022년.xlsx')
            ws = wb[f'{datetime.now().month}월']
                
            days = ['SUN', 'MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT']
            col = days.index(TODAY)
            
            # 셀 주소 찾기
            for row in ws.iter_rows(max_row=30, max_col=7):
                if row[col].value == datetime.now().day:
                    row_alphabet = row[col].coordinate[0]
                    col_num = int(row[col].coordinate[1:])
                    # 변수 = ws['{0}{1}'.format(row_alphabet, col_num + 4)] => 이렇게 하면 설정된 '변수'에 값을 할당해도 엑셀에 반영이 안 됨
            
            try: # json 파일 불러오기
                with open('time_data.json', 'r') as file:
                    time_data = json.load(file)
                    
            except FileNotFoundError: # json 파일 없으면 새로 만들어서 오늘 기록 저장
                new_record = {
                'record_of_the_day': self.time_unit.duration_of_the_day,
                'record_until_now': self.time_unit.duration_of_the_day,
                THIS_YEAR: {THIS_MONTH: self.time_unit.duration_of_the_day}
                }
                
                with open('time_data.json', 'w') as file:
                    json.dump(new_record, file, indent=4)
                
                # 여러 변수에 동일한 값 할당하기
                ws['{0}{1}'.format(row_alphabet, col_num + 4)] = ws['I5'] = ws['I7'] = self.time_unit.record_formatting(self.time_unit.duration_of_the_day)
            
            else: # json 파일 있으면 오늘 기록 추가하기              
                time_data['record_of_the_day'] = self.time_unit.duration_of_the_day
                time_data['record_until_now'] += self.time_unit.duration_of_the_day
                
                if THIS_YEAR not in time_data: # Check if 'the current year' exists in time data
                    time_data[THIS_YEAR] = {THIS_MONTH: self.time_unit.duration_of_the_day}
                
                elif THIS_MONTH in time_data[THIS_YEAR]: # Check if 'the current month' of 'the current year' exists in time data
                    time_data[THIS_YEAR][THIS_MONTH] += self.time_unit.duration_of_the_day
                    
                else:
                    time_data[THIS_YEAR][THIS_MONTH] = self.time_unit.duration_of_the_day
                    
                with open('time_data.json', 'w') as file:
                    json.dump(time_data, file, indent=4)
                
                # 오늘 기록이 해당 셀에 이미 있는 경우
                if ws['{0}{1}'.format(row_alphabet, col_num + 4)].value != None:
                    cell_hour, cell_min, cell_sec = map(int, ws['{0}{1}'.format(row_alphabet, col_num + 4)].value.split(':'))
                    cell_total_sec = (cell_hour * 3600) + (cell_min * 60) + cell_sec
                    result_sec = cell_total_sec + self.time_unit.duration_of_the_day
                    
                    ws['{0}{1}'.format(row_alphabet, col_num + 4)] = self.time_unit.record_formatting(result_sec)
                
                else: 
                    ws['{0}{1}'.format(row_alphabet, col_num + 4)] = self.time_unit.record_formatting(self.time_unit.duration_of_the_day)

                ws['I5'] = self.time_unit.record_formatting(time_data[THIS_YEAR][THIS_MONTH])
                ws['I7'] = self.time_unit.record_formatting(time_data['record_until_now'])
                
            finally: # tkinter 업데이트 & 엑셀 파일 저장
                with open('time_data.json', 'r') as file:
                    time_data = json.load(file)
                    self.this_month_label.config(text='this month {0}'.format(self.time_unit.record_formatting(time_data[THIS_YEAR][THIS_MONTH])))
                    self.until_now_label.config(text='until now {0}'.format(self.time_unit.record_formatting(time_data['record_until_now'])))
                    
                wb.save('성공2022년.xlsx')
                    
time_unit = TimeUnit()
studytimer = StudyTimer(time_unit)
